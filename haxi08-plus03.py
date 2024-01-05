import csv
import openpyxl
import xlrd
import os
import time
from preprocess_seq import seq_standardize
import random
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import get_column_letter
from collections import Counter

csv.field_size_limit(2147483647)


def geshi(filename):
    dataframe = pd.read_excel(filename)

    # 按照第二列进行排序
    dataframe_sorted = dataframe.sort_values(by=dataframe.columns[1])

    # 保存排序结果到新的 Excel 文件
    dataframe_sorted.to_excel(filename, index=False)


class MyStruct:
    def __init__(self, snp_code, gene_name, direction, gene_seq, snp_num):
        self.snp_code = snp_code
        self.gene_name = gene_name  # SNP信息
        self.direction = direction  # "left"或"right"，表示左侧序列或右侧序列
        self.gene_seq = gene_seq  # 匹配的子串
        self.snp_num = snp_num


def convert_to_binary_P(sequence):
    # 将传入的基因序列先转化为二进制，再转化为十进制  P代表正链
    binary = ''
    for char in sequence:
        if char == 'N':
            base_list = ['A', 'T', 'C', 'G']
            char = random.choice(base_list)
        if char == 'A':
            binary += '00'
        elif char == 'T':
            binary += '01'
        elif char == 'C':
            binary += '10'
        elif char == 'G':
            binary += '11'
        else:
            return None
    return int(binary, 2)


def convert_to_binary_N(sequence):
    # 将传入的基因序列先转化为二进制，再转化为十进制 N代表反链
    binary = ''
    for char in sequence:
        if char == 'N':
            base_list = ['A', 'T', 'C', 'G']
            char = random.choice(base_list)
        if char == 'T':
            binary += '00'
        elif char == 'A':
            binary += '01'
        elif char == 'G':
            binary += '10'
        elif char == 'C':
            binary += '11'
        else:
            return None
    return int(binary, 2)


def read_excel():
    # 处理snp库
    data = []
    workbook = xlrd.open_workbook("NewSNP.xls")
    sheet = workbook.sheet_by_index(0)
    for row in range(0, sheet.nrows):
        snp_code = sheet.cell_value(row, 0)
        snp_info = sheet.cell_value(row, 1)  # 获取SNP信息
        left_reads1 = sheet.cell_value(row, 2)  # 获取左侧序列
        right_reads2 = sheet.cell_value(row, 3)  # 获取右侧序列
        data.append((snp_code, snp_info, left_reads1, right_reads2))
    return data  # 返回读取的snp信息


def process_SNP():
    # 主处理函数

    snp_data = read_excel()  # 获取snp数据
    for snp_code, snp_info, left_reads1, right_reads2 in snp_data:
        index_1 = left_reads1.find("[")
        index_2 = right_reads2.find("[")
        index_3 = left_reads1.find("]")
        index_4 = right_reads2.find("]")
        snp_NUM[snp_info] = ""
        if index_1 != -1 and index_1 - match_num >= 0 and index_3 != -1 and len(left_reads1) - index_3 - match_num - 1 >= 0 :
            substring_1_qian = left_reads1[index_1 - match_num:index_1]
            substring_1_hou = left_reads1[index_3 + 1:index_3 + match_num + 1]
            temp_1=substring_1_qian+substring_1_hou
            # if snp_info=="FOXP1 rs17008544":
            #     print(temp_1)
            #     print("========")
            snp_binary_1 = convert_to_binary_P(temp_1)
            if snp_binary_1 in hash_table:
                res = MyStruct(snp_code, snp_info, "left_reads1", temp_1, 0)
                hash_table[snp_binary_1].append(res)
            else:
                hash_table[snp_binary_1] = []
                res = MyStruct(snp_code, snp_info, "left_reads1", temp_1, 0)
                hash_table[snp_binary_1].append(res)
        # if index_3 != -1 and len(left_reads1) - index_3 - match_num - 1 >= 0:
        #     substring_1 = left_reads1[index_3 + 1:index_3 + match_num + 1]
        #     snp_binary_1 = convert_to_binary_P(substring_1)
        #     if snp_binary_1 in hash_table:
        #         res = MyStruct(snp_code, snp_info, "left_reads1_后侧", substring_1, 0)
        #         hash_table[snp_binary_1].append(res)
        #     else:
        #         hash_table[snp_binary_1] = []
        #         res = MyStruct(snp_code, snp_info, "left_reads1_后侧", substring_1, 0)
        #         hash_table[snp_binary_1].append(res)
        if index_2 != -1 and index_2 - match_num >= 0 and index_4 != -1 and len(right_reads2) - index_4 - match_num - 1 >= 0:
            substring_2_qian = right_reads2[index_2 - match_num:index_2]
            substring_2_hou = right_reads2[index_4 + 1:index_4 + match_num + 1]
            temp_2=substring_2_qian+substring_2_hou
            snp_binary_2 = convert_to_binary_P(temp_2)
            if snp_binary_2 in hash_table:
                res = MyStruct(snp_code, snp_info, "right_reads2", temp_2, 0)
                hash_table[snp_binary_2].append(res)
            else:
                hash_table[snp_binary_2] = []
                res = MyStruct(snp_code, snp_info, "right_reads2", temp_2, 0)
                hash_table[snp_binary_2].append(res)
        # if index_4 != -1 and len(right_reads2) - index_4 - match_num - 1 >= 0:
        #     substring_1 = right_reads2[index_4 + 1:index_4 + match_num + 1]
        #     snp_binary_1 = convert_to_binary_P(substring_1)
        #     if snp_binary_1 in hash_table:
        #         res = MyStruct(snp_code, snp_info, "right_reads2_后侧", substring_1, 0)
        #         hash_table[snp_binary_1].append(res)
        #     else:
        #         hash_table[snp_binary_1] = []
        #         res = MyStruct(snp_code, snp_info, "right_reads2_后侧", substring_1, 0)
        #         hash_table[snp_binary_1].append(res)


def read_fasta_csv(file_path):
    with open(file_path, 'r') as file:
        reader = csv.reader(file)
        next(reader)  # 跳过标题行
        sequences = {}
        for row in reader:
            replacement = random.choice('ATCG')
            sequence_data = row[1].upper()  # 将序列数据转换为大写
            sequence_data = sequence_data.replace('N', replacement)
            sequences[row[0]] = sequence_data

    return sequences


def rep(sequence):  # 进行正反链转化
    sequence = sequence.replace('A', 'X')
    sequence = sequence.replace('T', 'A')
    sequence = sequence.replace('X', 'T')
    sequence = sequence.replace('C', 'Y')
    sequence = sequence.replace('G', 'C')
    sequence = sequence.replace('Y', 'G')

    return sequence


def match(file_path, filename):
    wb = Workbook()
    ws = wb.active
    ws.append(["snp_code", "snp_info", "序列方向", "匹配的序列", "id", "突变位置",
               "突变的碱基", "全基因组标记"])

    if os.path.isfile(file_path):
        seq = read_fasta_csv(file_path)
        for seq_id, seq_data in seq.items():
            # flag = Chain(seq_data)  # flag标记正反链
            sequence = seq_data  # 获取基因序列
            length = len(sequence)
            if length - 2*match_num - 3 >= 0:
                for i in range(match_num, length - match_num - 3, 1):
                    patch_1=sequence[i-match_num:i]
                    patch_2= sequence[i+1:i + match_num+1]
                    patch=patch_1+patch_2
                    # print(patch)
                    binary = convert_to_binary_P(patch)
                    if binary in hash_table:
                        for j in range(0, len(hash_table[binary]), 1):
                            if len(snp_NUM[hash_table[binary][j].gene_name]) < snp_limit:
                                if hash_table[binary][j].direction == "left_reads1":
                                    line = [hash_table[binary][j].snp_code, hash_table[binary][j].gene_name,
                                            hash_table[binary][j].direction,
                                            hash_table[binary][j].gene_seq, seq_id,
                                            i, sequence[i], Chain(sequence)]
                                    snp_NUM[hash_table[binary][j].gene_name] += sequence[i]
                                    ws.append(line)

                                # if hash_table[binary][j].direction == "left_reads1_后侧":
                                #     line = [hash_table[binary][j].snp_code, hash_table[binary][j].gene_name,
                                #             hash_table[binary][j].direction,
                                #             hash_table[binary][j].gene_seq, seq_id,
                                #             i - 1, sequence[i - 1], Chain(sequence)]
                                #     snp_NUM[hash_table[binary][j].gene_name] += sequence[i - 1]
                                #     ws.append(line)

                                if hash_table[binary][j].direction == "right_reads2":
                                    line = [hash_table[binary][j].snp_code, hash_table[binary][j].gene_name,
                                            hash_table[binary][j].direction,
                                            hash_table[binary][j].gene_seq, seq_id,
                                            i, rep(sequence[i]), Chain(sequence)]
                                    snp_NUM[hash_table[binary][j].gene_name] += rep(sequence[i])
                                    ws.append(line)

                                # if hash_table[binary][j].direction == "right_reads2_后侧":
                                #     line = [hash_table[binary][j].snp_code, hash_table[binary][j].gene_name,
                                #             hash_table[binary][j].direction,
                                #             hash_table[binary][j].gene_seq, seq_id,
                                #
                                #             i - 1, rep(sequence[i - 1]), Chain(sequence)]
                                #     snp_NUM[hash_table[binary][j].gene_name] += rep(sequence[i - 1])
                                #     ws.append(line)
    new_filename = filename[:-4]
    file = os.path.join('output', new_filename + '___table1.xlsx')
    wb.save(file)
    geshi(file)


def Chain(bef):  # 判断是正链还是反链
    if bef[-2:] == '/2':
        return '2(反)'
    elif bef[-2:] == '/1':
        return '1(正)'
    elif bef[-2:] == '/0':
        return '0(无)'


def process_table2(filename):
    src_workbook = xlrd.open_workbook('NewSNP.xls')
    src_worksheet = src_workbook.sheet_by_index(0)

    src_data = []
    for row in range(1, src_worksheet.nrows):  # 从第二行开始读取，跳过表头
        row_data = [cell.value for cell in src_worksheet.row(row)[:2]]
        src_data.append(row_data)

    # 创建新的 Excel 文件并写入数据
    dst_workbook = Workbook()
    dst_worksheet = dst_workbook.active  # 使用默认的工作表

    # 添加标题行
    headers = ['snp_code', 'snp_info', 'snp顺序结果', 'snp统计结果']
    for col_index, header_text in enumerate(headers):
        col_letter = get_column_letter(col_index + 1)
        dst_cell = dst_worksheet['{}1'.format(col_letter)]
        dst_cell.value = header_text

    for row_index, row_data in enumerate(src_data):
        for col_index, cell_data in enumerate(row_data):
            dst_cell = dst_worksheet.cell(row=row_index + 2, column=col_index + 1)
            dst_cell.value = cell_data

    # 保存并关闭 Excel 文件
    new_filename = filename[:-4]
    file = os.path.join('output', new_filename + '___table2.xlsx')
    dst_workbook.save(file)
    update_excel_with_hash_table(file)


def sort_bases(sequence):
    bases = {'A': False, 'T': False, 'G': False, 'C': False}
    result = []

    for base in sequence:
        if base in bases and not bases[base]:
            result.append(base)
            bases[base] = True

    order = {'A': 0, 'T': 1, 'G': 2, 'C': 3}
    result_sorted = sorted(result, key=lambda x: order[x])
    sorted_result_str = ''.join(result_sorted)

    base_counts = Counter(sequence)
    count_str_list = [f"{base}{base_counts[base]}" if base_counts[base] > 0 else "" for base in 'ATGC']
    count_str = ''.join(count_str_list)

    return sorted_result_str, count_str


def update_excel_with_hash_table(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        cell_value = row[0].value

        if len(snp_NUM[cell_value]) != 0:
            # 键存在于 hash_table 中，将值写入对应行的第三列
            value3, value4 = sort_bases(snp_NUM[cell_value])

            sheet.cell(row=row[0].row, column=3).value = value3
            sheet.cell(row=row[0].row, column=4).value = value4

        else:
            # 键不存在于 hash_table 中，对应行的第三列写入空
            sheet.cell(row=row[0].row, column=3).value = "0"
            sheet.cell(row=row[0].row, column=4).value = "A0T0G0C0"

    wb.save(excel_file_path)


def process_table3():
    workbook1 = xlrd.open_workbook("NewSNP.xls")
    sheet1 = workbook1.sheet_by_index(0)

    # 创建新的Excel文件，用于保存复制后的数据
    workbook2 = Workbook()
    sheet2 = workbook2.active
    sheet2.title = 'Sheet2'
    # 从原始表格中复制前两列数据到新表格中
    for row_idx in range(sheet1.nrows):
        row_data = sheet1.row_values(row_idx, end_colx=2)  # 读取指定列范围的数据
        sheet2.append(row_data)

    # 保存新表格
    i = 3
    file_directory = "output"

    for file in os.listdir(file_directory):
        if file.endswith('___table2.xlsx'):
            top_name = file[:-14]
            file_path = os.path.join(file_directory, file)
            workbook = load_workbook(file_path)

            # 获取第一个工作表
            sheet1 = workbook.active

            # 遍历 sheet1 中的每一行
            for row in sheet1.iter_rows():
                # 读取第三列的数据
                value = row[2].value

                # 将数据添加到 sheet2 的对应列
                sheet2.cell(row=row[0].row, column=i, value=value)
            sheet2.cell(row=1, column=i, value=top_name)

            i += 1

    # 保存新的工作簿
    file = os.path.join('output', 'table3.xlsx')
    workbook2.save(file)


if __name__ == '__main__':
    start_time = time.time()
    match_num = int(input("输入 匹配长度"))
    snp_limit = int(input("输入 一条snp的匹配数量限制"))
    folder_name = input("输入文件夹名称地址")
    # folder_name = "big_loop/"
    seq_standardize(folder_name)

    snp_NUM = {}
    hash_table = {}
    print("运行中")
    for filename in os.listdir("data_std"):
        process_SNP()
        file_path = os.path.join("data_std", filename)
        match(file_path, filename)
        process_table2(filename)
        # print("===")
        snp_NUM = {}
        hash_table = {}
    process_table3()
    end_time = time.time()
    print("运行结束，时间为:  ", end_time - start_time, "秒")
